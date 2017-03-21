#!usr/bin/env python
# -*- coding: utf-8 -*-
from plugin_notemptymarker import mark_not_empty
import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font
import unicodedata
import calendar


@auth.requires(auth.has_membership('Director') \
  or auth.has_membership('Gestor de Sustancias') \
  or auth.has_membership('WebMaster'))
@auth.requires_login()

def select_rl4():
	#lab = str(db(db.t_laboratorio.id == request.vars['lab']).select(db.t_laboratorio.f_nombre))[24:-2]
    #query = (db.t_sustancias.f_control == 3) |  (db.t_sustancias.f_control == 1)
    mes = request.vars['m']
    year= request.vars['y']
    query=(db.v_reporte.f_mes==mes)&(db.v_reporte.f_year==year)
    query2=db(db.t_bitacora.f_sustancia==2).select()
    for i in query2:
    	print (i.f_fechaingreso)
    table = SQLFORM.smartgrid(db.v_reporte,constraints=dict(v_reporte=query),csv=False,editable=False,deletable=False,create=False)
    return locals()

def select_fecha():
	now = datetime.datetime.now()
	table = SQLFORM.factory(Field('mes',requires=IS_IN_SET(['Enero','Febrero','Marzo','Abril','Mayo','Junio','Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']),
			label=T('Seleccione mes')),
		Field('year','integer',requires=IS_INT_IN_RANGE(1969,now.year+1,error_message='Debe introducir un año menor o igual al actual'),
			label=T('Introduzca año'))
		)
	if table.process().accepted:
		if table.vars.mes=="Enero":
			x=1
		elif table.vars.mes=="Febrero":
			x=2
		elif table.vars.mes=="Marzo":
			x=3
		elif table.vars.mes=="Abril":
			x=4
		elif table.vars.mes=="Mayo":
			x=5
		elif table.vars.mes=="Junio":
			x=6
		elif table.vars.mes=="Julio":
			x=7
		elif table.vars.mes=="Agosto":
			x=8
		elif table.vars.mes=="Septiembre":
			x=9
		elif table.vars.mes=="Octubre":
			x=10
		elif table.vars.mes=="Noviembre":
			x=11
		elif table.vars.mes=="Diciembre":
			x=12
		redirect(URL('select_rl4',vars=dict(m=x,y=table.vars.year)))
	return locals()	


def generar_reporte():
	wb = Workbook()
	ws = wb.active
	cen = Alignment(horizontal='center', vertical='distributed')
	rig = Alignment(horizontal='right')
	lef = Alignment(horizontal='left')
	ft1 = Font(name='Arial', size=10, bold=True)
	ft2 = Font(name='Arial', size=10, bold=False)
	ft3 = Font(name='Arial', size=8)
	ws.font = ft2
	now = datetime.datetime.now()

	ws1 = wb.create_sheet("Informe mensual")

	#Encabezado
	ws.title = "Informe mensual"
	ws1.title = "Informe mensual"
	img = Image('gob.jpg')
	img1 = Image('gob.jpg')
	ws.add_image(img, 'A1')
	ws1.add_image(img1, 'A1')

	#tamaño de las columnas
	for i in ['A', 'D', 'E','F','G','J','K']:
	   ws.column_dimensions[i].width = 10
	   ws1.column_dimensions[i].width = 10
	ws.column_dimensions['B'].width = 17
	ws.column_dimensions['C'].width = 11
	ws.column_dimensions['H'].width = 9
	ws.column_dimensions['I'].width = 10
	ws1.column_dimensions['B'].width = 17
	ws1.column_dimensions['C'].width = 11
	ws1.column_dimensions['H'].width = 9
	ws1.column_dimensions['I'].width = 10

	#tamaño de las filas
	ws.row_dimensions[13].height = 40
	ws1.row_dimensions[13].height = 40
	for i in range(1,13):
	    ws.row_dimensions[i].height = 13
	    ws1.row_dimensions[i].height = 13
	for i in range(14,29):
	    ws.row_dimensions[i].height = 12
	    ws1.row_dimensions[i].height = 12

	#All Merges
	ws.merge_cells(start_row=5,start_column=3,end_row=5,end_column=10)
	ws.merge_cells(start_row=7,start_column=3,end_row=7,end_column=5)
	ws1.merge_cells(start_row=5,start_column=3,end_row=5,end_column=10)
	ws1.merge_cells(start_row=7,start_column=3,end_row=7,end_column=5)
	for i in range(13,28):
	    ws.merge_cells(start_row=i,start_column=2,end_row=i,end_column=3)
	    ws.merge_cells(start_row=i,start_column=10,end_row=i,end_column=11)
	    ws1.merge_cells(start_row=i,start_column=2,end_row=i,end_column=3)
	    ws1.merge_cells(start_row=i,start_column=10,end_row=i,end_column=11)

	for i in range(29,33):
	    ws.merge_cells(start_row=i,start_column=1,end_row=i,end_column=10)
	    ws1.merge_cells(start_row=i,start_column=1,end_row=i,end_column=10)

	#titulos y datos
	z = ['C5', 'J7', 'I9', 'J9', 'K9','B7','B8','B9','B10','B11','I10','J10','K10']
	ws['C5'] = 'INFORME MENSUAL DE SUSTANCIAS QUIMICAS CONTROLADAS'
	ws['J7'] = 'FECHA'
	ws['I9'] = 'DIA'
	ws['J9'] = 'MES'
	ws['K9'] = 'AÑO'
	ws1['C5'] = 'INFORME MENSUAL DE SUSTANCIAS QUIMICAS CONTROLADAS'
	ws1['J7'] = 'FECHA'
	ws1['I9'] = 'DIA'
	ws1['J9'] = 'MES'
	ws1['K9'] = 'AÑO'


	for i in range(5):
	    ws[z[i]].font = ft1
	    ws[z[i]].alignment = cen
	    ws1[z[i]].font = ft1
	    ws1[z[i]].alignment = cen

	ws['B7'] = 'OPERADOR:'
	ws['B8'] = 'LICENCIA:'
	ws['B9'] = 'PERMISO DEL CICPC:'
	ws['B10'] = 'RIF:'
	ws['B11'] = 'MES-AÑO:'
	ws1['B7'] = 'OPERADOR:'
	ws1['B8'] = 'LICENCIA:'
	ws1['B9'] = 'PERMISO DEL CICPC:'
	ws1['B10'] = 'RIF:'
	ws1['B11'] = 'MES-AÑO:'

	for i in range(5,10):
	    ws[z[i]].font = ft1
	    ws[z[i]].alignment = rig
	    ws1[z[i]].font = ft1
	    ws1[z[i]].alignment = rig

	ws['I10'] = now.day
	ws['J10'] = now.month
	ws['K10'] = now.year
	ws1['I10'] = now.day
	ws1['J10'] = now.month
	ws1['K10'] = now.year

	for i in range(10,13):
	    ws[z[i]].font = ft2
	    ws[z[i]].alignment = cen
	    ws1[z[i]].font = ft2
	    ws1[z[i]].alignment = cen

	ws['A28'] = 'Nota:'
	ws['A28'].font = ft1
	ws['A28'].alignment = lef
	ws1['A28'] = 'Nota:'
	ws1['A28'].font = ft1
	ws1['A28'].alignment = lef

	w = ['C7', 'C8', 'C9', 'C10', 'C11','A13','B13','D13','E13','F13','G13','H13','I13','J13']

	for i in range(5):
	    ws[w[i]].font = ft2
	    ws1[w[i]].font = ft2

	ws['C7'] = 'UNIVERSIDAD SIMON BOLIVAR'
	ws['C8'] = '2014LIC0256'
	ws['C9'] = 'No. 1311'
	ws['C10'] = 'G-20000063-5'
	ws['C11'] = str(now.month)+'/'+str(now.year)
	ws['A13'] = 'N°'

	ws['B13'] = 'Sustancia Química Controlada'

	ws['D13'] = 'Código Arancelario'

	ws['E13'] = 'Saldo Físico Inicial'

	ws['F13'] = 'Total Entradas'

	ws['G13'] = 'Total Salidas'

	ws['H13'] = 'Saldo Físico Final'

	ws['I13'] = 'Unidad de Medida'

	ws['J13'] = 'Observaciones'

	ws1['C7'] = 'UNIVERSIDAD SIMON BOLIVAR'
	ws1['C8'] = '2014LIC0256'
	ws1['C9'] = 'No. 1311'
	ws1['C10'] = 'G-20000063-5'
	ws1['C11'] = str(now.month)+'/'+str(now.year)
	ws1['A13'] = 'N°'

	ws1['B13'] = 'Sustancia Química Controlada'

	ws1['D13'] = 'Código Arancelario'

	ws1['E13'] = 'Saldo Físico Inicial'

	ws1['F13'] = 'Total Entradas'

	ws1['G13'] = 'Total Salidas'

	ws1['H13'] = 'Saldo Físico Final'

	ws1['I13'] = 'Unidad de Medida'

	ws1['J13'] = 'Observaciones'


	for i in range(5,14):
	    ws[w[i]].font = ft1
	    ws[w[i]].alignment = cen
	    ws1[w[i]].font = ft1
	    ws1[w[i]].alignment = cen


	x = ['A14','A15','A16','A17','A18','A19','A20','A21','A22','A23','A24','A25','A26']
	y = ['01','02','03','04','05','06','07','08','09','10','11','12','13']
	for i in range(0,13):
	    ws[x[i]] = y[i]
	    ws[x[i]].font = ft3
	    ws[x[i]].alignment = cen
	    ws1[x[i]] = y[i]
	    ws1[x[i]].font = ft3
	    ws1[x[i]].alignment = cen
	x = ['B14','B15','B16','B17','B18','B19','B20','B21','B22','B23','B24','B25','B26']
	mes = request.vars['m']
	year= request.vars['y']
	query=db((db.v_reporte.f_mes==mes)&(db.v_reporte.f_year==year)).select(db.v_reporte.f_nombre)
	y=0;
	z=0;
	for i in query:
		if y<13:
			ws[x[y]] = i.f_nombre
			ws[x[y]].font = ft3
	    	y=y+1
		if y>=13:
			ws1[x[z]] = i.f_nombre
			ws1[x[z]].font = ft3
			z=z+1
	x = ['E14','E15','E16','E17','E18','E19','E20','E21','E22','E23','E24','E25','E26']
	query=db((db.v_reporte.f_mes==mes)&(db.v_reporte.f_year==year)).select(db.v_reporte.f_nombre)
	y=0;
	z=0;
	for i in query:
		if y<13:
			print("Hice el query")
			query1=db((db.v_reporte.f_mes<int(mes))&(db.v_reporte.f_year==int(year))&(db.v_reporte.f_nombre==i.f_nombre)).select(db.v_reporte.cantidad_total)
			aux1=0
			entro=0
			for p in query1:
				if aux1==0:
					print("ENTRE AQUI")
					print(p.cantidad_total/1000,i.f_nombre)
					ws[x[y]] = (p.cantidad_total/1000)
					ws[x[y]].font = ft3
					aux1=aux1+1
					entro=entro+1
					y=y+1
			print (entro)		
			if entro==0:
				query2=db((db.v_reporte.f_year<int(year))&(db.v_reporte.f_nombre==i.f_nombre)).select(db.v_reporte.cantidad_total)
				aux=0
				for u in query2:
					if aux==0:
						ws[x[y]] = (u.cantidad_total/1000)
						ws[x[y]].font = ft3
						aux=aux+1
						entro=entro+1
						y=y+1
			if entro==0:
				ws[x[y]] = 0
				ws[x[y]].font = ft3
				y=y+1



		if y>=13:		
			print("Hice el query")
			query1=db((db.v_reporte.f_mes<mes)&(db.v_reporte.f_year==year)&(db.v_reporte.f_nombre==i.f_nombre)).select(db.v_reporte.cantidad_total)
			aux1=0
			entro=0
			for p in query1:
				if aux1==0:
					ws1[x[y]] = (u.cantidad_total/1000)
					ws1[x[y]].font = ft3
					aux1=aux1+1
					entro=entro+1
			if entro==0:
				query2=db((db.v_reporte.f_year<year)&(db.v_reporte.f_nombre==i.f_nombre)).select(db.v_reporte.cantidad_total)
				aux=0
				for u in query2:
					if aux==0:
						ws1[x[y]] = (u.cantidad_total/1000)
						ws1[x[y]].font = ft3
						aux=aux+1
						entro=entro+1
			if entro==0:
				ws1[x[y]] = 0
				ws1[x[y]].font = ft3
	x = ['F14','F15','F16','F17','F18','F19','F20','F21','F22','F23','F24','F25','F26']
	query=db((db.v_reporte.f_mes==mes)&(db.v_reporte.f_year==year)).select(db.v_reporte.total_entradas)
	y=0;
	z=0;
	for i in query:
		if y<13:
			ws[x[y]] = (i.total_entradas/1000)
			ws[x[y]].font = ft3
	    	y=y+1
		if y>=13:
			ws1[x[y]] = (i.total_entradas/1000)
			ws1[x[y]].font = ft3
			y=y+1
	x = ['G14','G15','G16','G17','G18','G19','G20','G21','G22','G23','G24','G25','G26']
	query=db((db.v_reporte.f_mes==mes)&(db.v_reporte.f_year==year)).select(db.v_reporte.total_salidas)
	y=0;
	z=0;
	for i in query:
		if y<13:
			ws[x[y]] = (i.total_salidas/1000)
			ws[x[y]].font = ft3
	    	y=y+1
		if y>=13:
			ws1[x[z]] = (i.total_salidas/1000)
			ws1[x[z]].font = ft3
			z=z+1
	x = ['H14','H15','H16','H17','H18','H19','H20','H21','H22','H23','H24','H25','H26']
	query=db((db.v_reporte.f_mes==mes)&(db.v_reporte.f_year==year)).select(db.v_reporte.cantidad_total)
	y=0;
	z=0;
	for i in query:
		if y<13:
			ws[x[y]] = (i.cantidad_total/1000)
			ws[x[y]].font = ft3
	    	y=y+1
		if y>=13:
			ws1[x[z]] = (i.cantidad_total/1000)
			ws1[x[z]].font = ft3
			z=z+1
	x = ['I14','I15','I16','I17','I18','I19','I20','I21','I22','I23','I24','I25','I26']
	query=db((db.v_reporte.f_mes==mes)&(db.v_reporte.f_year==year)).select(db.v_reporte.f_unidad)
	y=0;
	z=0;
	for i in query:
		if y<13:
			if i.f_unidad=="mL":
				ws[x[y]] = "L"
			ws[x[y]].font = ft3
			if i.f_unidad=="g":
				ws[x[y]] = "Kg"
	    	y=y+1
		if y>=13:
			ws1[x[z]] = i.f_unidad
			ws1[x[z]].font = ft3
			z=z+1
	#ws3 = wb["New Title"]

	#Pie de Pagina
	ws['A29'] = '1. Los saldos serán reportados en:'
	ws['A30'] = 'Kgs. Para sustancias en estado sólido ó Lts. Para sustancias en estado líquido, especificando la densidad de la sustancia en el último caso.'
	ws['A31'] = '2. El reporte mensual será llevado por cada sustancia química controlada'
	ws['A32'] = '3. El reporte mensual deberá ser entregado dentro de los primeros 7 días hábiles de cada mes'
	ws1['A29'] = '1. Los saldos serán reportados en:'
	ws1['A30'] = 'Kgs. Para sustancias en estado sólido ó Lts. Para sustancias en estado líquido, especificando la densidad de la sustancia en el último caso.'
	ws1['A31'] = '2. El reporte mensual será llevado por cada sustancia química controlada'
	ws1['A32'] = '3. El reporte mensual deberá ser entregado dentro de los primeros 7 días hábiles de cada mes'
	#wb.save('Reporte Universidad Simon Bolivar.xlsx')


	#Aqui comenzamos con los reportes individuales
	nombres=[]
	query=db((db.v_reporte.f_mes==mes)&(db.v_reporte.f_year==year)).select(db.v_reporte.f_nombre)
	for i in query:
		nombres.append(i.f_nombre)
	ids=[]
	for i in nombres:
		query2=db(db.t_sustancias.f_nombre==i).select(db.t_sustancias.id)
		for t in query2:
			ids.append(t.id)
	bitacora=[]
	query5=db((db.t_bitacora.f_fechaingreso.month==1)).select(db.t_bitacora.f_fechaingreso)
	for h in query5:
		print(h.f_fechaingreso)
	for y in ids:
		print(mes,year)
		query3=db((db.t_bitacora.f_sustancia==y)).select()
		aux=0;
		for z in query3:
			if(z.f_fechaingreso.month==int(mes) and z.f_fechaingreso.year==int(year)):
				aux=aux+1
		bitacora.append(aux)
	print(bitacora)
	print(ids)

	contador=0
	for m in nombres:
		hoja=1;
		contador=contador+1
		while hoja<bitacora[contador-1]:
			m=unicode(m,"utf-8")
			ws2 = wb.create_sheet(m)
			#Encabezado
			ws2.title = m
			img = Image('gob.jpg')
			ws2.add_image(img, 'A1')


			#tamaño de las columnas
			for i in ['A', 'D', 'E','F','J','K']:
			   ws2.column_dimensions[i].width = 9
			ws2.column_dimensions['B'].width = 9
			ws2.column_dimensions['C'].width = 17.5
			ws2.column_dimensions['H'].width = 9
			ws2.column_dimensions['I'].width = 10.5
			ws2.column_dimensions['G'].width = 11

			#tamaño de las filas
			ws2.row_dimensions[13].height = 13
			for i in range(1,13):
			    ws2.row_dimensions[i].height = 13
			for i in range(14,42):
			    ws2.row_dimensions[i].height = 13


			#All Merges
			ws2.merge_cells(start_row=5,start_column=2,end_row=5,end_column=7)
			ws2.merge_cells(start_row=7,start_column=3,end_row=7,end_column=5)

			#titulos y datos
			z = ['B5', 'G7', 'F8', 'G8', 'H8','B7','B8','B9','B10','B11','B12','F9','G9','H9']
			ws2['B5'] = 'INFORME DE REPORTE DIARIO DE SUSTANCIAS QUIMICAS CONTROLADAS'
			ws2['G7'] = 'FECHA'
			ws2['F8'] = 'PERIODO'
			ws2['G8'] = 'MES'
			ws2['H8'] = 'AÑO'


			for i in range(5):
			    ws2[z[i]].font = ft1
			    ws2[z[i]].alignment = cen


			ws2['B7'] = 'OPERADOR:'
			ws2['B8'] = 'LICENCIA:'
			ws2['B9'] = 'PERMISO DEL CICPC:'
			ws2['B10'] = 'RIF:'
			ws2['B11'] = 'SUSTANCIA:'
			ws2['B12'] = 'UNIDAD'


			for i in range(5,11):
			    ws2[z[i]].font = ft1
			    ws2[z[i]].alignment = rig

			ws2['F9'] = '1 al '+str(calendar.monthrange(now.year,now.month)[1])
			ws2['G9'] = now.month
			ws2['H9'] = now.year

			for i in range(11,14):
			    ws2[z[i]].font = ft2
			    ws2[z[i]].alignment = cen


			ws2['A36'] = 'Nota:'
			ws2['A36'].font = ft1
			ws2['A36'].alignment = lef


			w = ['C7', 'C8', 'C9', 'C10', 'C11','A14','B14','C14','D14','E14','F14','G14']

			for i in range(5):
			    ws2[w[i]].font = ft2

			ws2['C7'] = 'UNIVERSIDAD SIMON BOLIVAR'
			ws2['C8'] = '2014LIC0256'
			ws2['C9'] = 'No. 1311'
			ws2['C10'] = 'G-20000063-5'
			ws2['C11'] = m.upper()
			query4=db(db.v_reporte.f_nombre==m).select(db.v_reporte.f_unidad)
			for l in query4:
				print(l.f_unidad)
				if l.f_unidad=="mL":
					l.f_unidad="L"
				if l.f_unidad=="g":
					l.f_unidad="Kg"
				if l.f_unidad==None:
					l.f_unidad="Nada"
			
			ws2['C12']=l.f_unidad 
			ws2['C12'].font = ft2
			ws2['C12'].alignment = cen 			
			
			ws2['A14'] = 'Asiento'

			ws2['B14'] = 'Fecha'

			ws2['C14'] = 'Descripcion del proceso'

			ws2['D14'] = 'Entrada'

			ws2['E14'] = 'Salida'

			ws2['F14'] = 'Saldo Final'

			ws2['G14'] = 'Observaciones'

			for i in range(5,12):
			    ws2[w[i]].font = ft1
			    ws2[w[i]].alignment = cen

			x = ['A15','A16','A17','A18','A19','A20','A21','A22','A23','A24','A25','A26','A27','A28','A29','A30','A31','A32','A33','A34']
			y = ['01','02','03','04','05','06','07','08','09','10','11','12','13','14','15','16','17','18','19','20']
			for i in range(0,20):
				ws2[x[i]] = y[i]
				ws2[x[i]].font = ft2
				ws2[x[i]].alignment = cen

			x = ['B15','B16','B17','B18','B19','B20','B21','B22','B23','B24','B25','B26','B27','B28','B29','B30','B31','B32','B33','B34']
			if bitacora[contador-1]!=0:
				query=db(db.t_bitacora.f_sustancia==ids[contador-1]).select(orderby=db.t_bitacora.f_fechaingreso)
				y=0;
				for i in query:
					if y<20:
						if(i.f_fechaingreso.month==int(mes) and i.f_fechaingreso.year==int(year)):
							ws2[x[y]]= i.f_fechaingreso
							ws2[x[y]].font= ft3
							ws2[x[y]].alignment = cen
							y=y+1

			x = ['C15','C16','C17','C18','C19','C20','C21','C22','C23','C24','C25','C26','C27','C28','C29','C30','C31','C32','C33','C34']
			if bitacora[contador-1]!=0:
				query=db(db.t_bitacora.f_sustancia==ids[contador-1]).select(orderby=db.t_bitacora.f_fechaingreso)
				y=0;
				for i in query:
					if y<20:
						if(i.f_fechaingreso.month==int(mes) and i.f_fechaingreso.year==int(year)):
							ws2[x[y]]= i.f_proceso
							ws2[x[y]].font= ft3
							ws2[x[y]].alignment = cen
							y=y+1
			x = ['D15','D16','D17','D18','D19','D20','D21','D22','D23','D24','D25','D26','D27','D28','D29','D30','D31','D32','D33','D34']
			if bitacora[contador-1]!=0:
				query=db(db.t_bitacora.f_sustancia==ids[contador-1]).select(orderby=db.t_bitacora.f_fechaingreso)
				y=0;
				for i in query:
					if y<20:
						if(i.f_fechaingreso.month==int(mes) and i.f_fechaingreso.year==int(year)):
							ws2[x[y]]= i.f_ingreso/1000
							ws2[x[y]].font= ft3
							ws2[x[y]].alignment = cen
							y=y+1
			x = ['E15','E16','E17','E18','E19','E20','E21','E22','E23','E24','E25','E26','E27','E28','E29','E30','E31','E32','E33','E34']
			if bitacora[contador-1]!=0:
				query=db(db.t_bitacora.f_sustancia==ids[contador-1]).select(orderby=db.t_bitacora.f_fechaingreso)
				y=0;
				for i in query:
					if y<20:
						if(i.f_fechaingreso.month==int(mes) and i.f_fechaingreso.year==int(year)):
							ws2[x[y]]= i.f_consumo/1000
							ws2[x[y]].font= ft3
							ws2[x[y]].alignment = cen
							y=y+1
			x = ['F15','F16','F17','F18','F19','F20','F21','F22','F23','F24','F25','F26','F27','F28','F29','F30','F31','F32','F33','F34']
			if bitacora[contador-1]!=0:
				query=db(db.t_bitacora.f_sustancia==ids[contador-1]).select(orderby=db.t_bitacora.f_fechaingreso)
				y=0;
				for i in query:
					if y<20:
						if(i.f_fechaingreso.month==int(mes) and i.f_fechaingreso.year==int(year)):
							hoja=hoja+1
							ws2[x[y]]= i.f_cantidad/1000
							ws2[x[y]].font= ft3
							ws2[x[y]].alignment = cen
							y=y+1

			#Pie de Pagina
			ws2['A37'] = '1. Los saldos serán reportados en:'
			ws2['A38'] = 'Kgs. Para sustancias en estado sólido ó Lts. Para sustancias en estado líquido, especificando la densidad de la sustancia en el último caso.'
			ws2['A39'] = '2. El reporte mensual será llevado por cada sustancia química controlada'
			ws2['A40'] = '3. El reporte mensual deberá ser entregado dentro de los primeros 7 días hábiles de cada mes'

	wb.save('Reporte Universidad Simon Bolivar.xlsx')
	return locals()


def select_rl7():
	#lab = str(db(db.t_laboratorio.id == request.vars['lab']).select(db.t_laboratorio.f_nombre))[24:-2]
    #query = (db.t_sustancias.f_control == 3) |  (db.t_sustancias.f_control == 1)
    mes = request.vars['m']
    year= request.vars['y']
    query=(db.v_reporte_rl7.f_mes==mes)&(db.v_reporte_rl7.f_year==year)

    table = SQLFORM.smartgrid(db.v_reporte_rl7,constraints=dict(v_reporte_rl7=query),csv=False,editable=False,deletable=False,create=False)
    return locals()


	
def select_fecha7():
	now = datetime.datetime.now()
	table = SQLFORM.factory(Field('mes',requires=IS_IN_SET(['Enero','Febrero','Marzo','Abril','Mayo','Junio','Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']),
			label=T('Seleccione mes')),
		Field('year','integer',requires=IS_INT_IN_RANGE(1969,now.year+1,error_message='Debe introducir un año menor o igual al actual'),
			label=T('Introduzca año'))
		)
	if table.process().accepted:
		if table.vars.mes=="Enero":
			x=1
		elif table.vars.mes=="Febrero":
			x=2
		elif table.vars.mes=="Marzo":
			x=3
		elif table.vars.mes=="Abril":
			x=4
		elif table.vars.mes=="Mayo":
			x=5
		elif table.vars.mes=="Junio":
			x=6
		elif table.vars.mes=="Julio":
			x=7
		elif table.vars.mes=="Agosto":
			x=8
		elif table.vars.mes=="Septiembre":
			x=9
		elif table.vars.mes=="Octubre":
			x=10
		elif table.vars.mes=="Noviembre":
			x=11
		elif table.vars.mes=="Diciembre":
			x=12
		redirect(URL('select_rl7',vars=dict(m=x,y=table.vars.year)))
	return locals()	

def generar_reporte_rl7():
	mes = request.vars['m']
	year= request.vars['y']
	wb = Workbook()
	ws2 = wb.active
	cen = Alignment(horizontal='center', vertical='distributed')
	rig = Alignment(horizontal='right')
	lef = Alignment(horizontal='left')
	ft1 = Font(name='Arial', size=10, bold=True)
	ft2 = Font(name='Arial', size=10, bold=False)
	ft3 = Font(name='Arial', size=8)
	ws2.font = ft2
	now = datetime.datetime.now()
	nombres=[]
	query=db((db.v_reporte_rl7.f_mes==mes)&(db.v_reporte_rl7.f_year==year)).select(db.v_reporte_rl7.f_nombre)
	for i in query:
		nombres.append(i.f_nombre)
	ids=[]
	for i in nombres:
		query2=db(db.t_sustancias.f_nombre==i).select(db.t_sustancias.id)
		for t in query2:
			ids.append(t.id)
	bitacora=[]
	query5=db((db.t_bitacora.f_fechaingreso.month==1)).select(db.t_bitacora.f_fechaingreso)
	for h in query5:
		print(h.f_fechaingreso)
	for y in ids:
		print(mes,year)
		query3=db((db.t_bitacora.f_sustancia==y)).select()
		aux=0;
		for z in query3:
			if(z.f_fechaingreso.month==int(mes) and z.f_fechaingreso.year==int(year)):
				aux=aux+1
		bitacora.append(aux)
	print(nombres)
	print(bitacora)

	contador=0
	for m in nombres:
		hoja=1;
		contador=contador+1
		while hoja<=bitacora[contador-1]:
			m=unicode(m,"utf-8")
			ws2 = wb.create_sheet(m)
			#Encabezado
			ws2.title = m
			img = Image('gob.jpg')
			ws2.add_image(img, 'A1')


			#tamaño de las columnas
			for i in ['A', 'D', 'E','F','J','K']:
			   ws2.column_dimensions[i].width = 9
			ws2.column_dimensions['B'].width = 9
			ws2.column_dimensions['C'].width = 17.5
			ws2.column_dimensions['H'].width = 9
			ws2.column_dimensions['I'].width = 10.5
			ws2.column_dimensions['G'].width = 11

			#tamaño de las filas
			ws2.row_dimensions[13].height = 13
			for i in range(1,13):
			    ws2.row_dimensions[i].height = 13
			for i in range(14,42):
			    ws2.row_dimensions[i].height = 13


			#All Merges
			ws2.merge_cells(start_row=5,start_column=2,end_row=5,end_column=7)
			ws2.merge_cells(start_row=7,start_column=3,end_row=7,end_column=5)

			#titulos y datos
			z = ['B5', 'G7', 'F8', 'G8', 'H8','B7','B8','B9','B10','B11','B12','F9','G9','H9']
			ws2['B5'] = 'INFORME DE REPORTE DIARIO DE SUSTANCIAS QUIMICAS RL7'
			ws2['G7'] = 'FECHA'
			ws2['F8'] = 'PERIODO'
			ws2['G8'] = 'MES'
			ws2['H8'] = 'AÑO'


			for i in range(5):
			    ws2[z[i]].font = ft1
			    ws2[z[i]].alignment = cen


			ws2['B7'] = 'OPERADOR:'
			ws2['B8'] = 'LICENCIA:'
			ws2['B9'] = 'PERMISO DEL CICPC:'
			ws2['B10'] = 'RIF:'
			ws2['B11'] = 'SUSTANCIA:'
			ws2['B12'] = 'UNIDAD'


			for i in range(5,11):
			    ws2[z[i]].font = ft1
			    ws2[z[i]].alignment = rig

			ws2['F9'] = '1 al '+str(calendar.monthrange(now.year,now.month)[1])
			ws2['G9'] = now.month
			ws2['H9'] = now.year

			for i in range(11,14):
			    ws2[z[i]].font = ft2
			    ws2[z[i]].alignment = cen

			w = ['C7', 'C8', 'C9', 'C10', 'C11','A14','B14','C14','D14','E14','F14','G14']

			for i in range(5):
			    ws2[w[i]].font = ft2

			ws2['C7'] = 'UNIVERSIDAD SIMON BOLIVAR'
			ws2['C8'] = '2014LIC0256'
			ws2['C9'] = 'No. 1311'
			ws2['C10'] = 'G-20000063-5'
			ws2['C11'] = m.upper()
			query4=db(db.v_reporte_rl7.f_nombre==m).select(db.v_reporte_rl7.f_unidad)
			for l in query4:
				print(l.f_unidad)
				if l.f_unidad=="mL":
					l.f_unidad="L"
				if l.f_unidad=="g":
					l.f_unidad="Kg"
				if l.f_unidad==None:
					l.f_unidad="Nada"
			
			ws2['C12']=l.f_unidad 
			ws2['C12'].font = ft2
			ws2['C12'].alignment = cen 			
			
			ws2['A14'] = 'N°'

			ws2['B14'] = 'Fecha'

			ws2['C14'] = 'Descripcion del proceso'

			ws2['D14'] = 'Entrada'

			ws2['E14'] = 'Salida'

			ws2['F14'] = 'Saldo Final'

			ws2['G14'] = 'Observaciones'

			for i in range(5,12):
			    ws2[w[i]].font = ft1
			    ws2[w[i]].alignment = cen

			x = ['A15','A16','A17','A18','A19','A20','A21','A22','A23','A24','A25','A26','A27','A28','A29','A30','A31','A32','A33','A34','A35','A36','A37','A38']
			y = ['01','02','03','04','05','06','07','08','09','10','11','12','13','14','15','16','17','18','19','20','21','22','23','24']
			for i in range(0,24):
				ws2[x[i]] = y[i]
				ws2[x[i]].font = ft2
				ws2[x[i]].alignment = cen

			x = ['B15','B16','B17','B18','B19','B20','B21','B22','B23','B24','B25','B26','B27','B28','B29','B30','B31','B32','B33','B34','B35','B36','B37','B38']
			if bitacora[contador-1]!=0:
				query=db(db.t_bitacora.f_sustancia==ids[contador-1]).select(orderby=db.t_bitacora.f_fechaingreso)
				y=0;
				for i in query:
					if y<24:
						if(i.f_fechaingreso.month==int(mes) and i.f_fechaingreso.year==int(year)):
							ws2[x[y]]= i.f_fechaingreso
							ws2[x[y]].font= ft3
							ws2[x[y]].alignment = cen
							y=y+1

			x = ['C15','C16','C17','C18','C19','C20','C21','C22','C23','C24','C25','C26','C27','C28','C29','C30','C31','C32','C33','C34','AC5','C36','C37','C38']
			if bitacora[contador-1]!=0:
				query=db(db.t_bitacora.f_sustancia==ids[contador-1]).select(orderby=db.t_bitacora.f_fechaingreso)
				y=0;
				for i in query:
					if y<24:
						if(i.f_fechaingreso.month==int(mes) and i.f_fechaingreso.year==int(year)):
							ws2[x[y]]= i.f_proceso
							ws2[x[y]].font= ft3
							ws2[x[y]].alignment = cen
							y=y+1
			x = ['D15','D16','D17','D18','D19','D20','D21','D22','D23','D24','D25','D26','D27','D28','D29','D30','D31','D32','D33','D34','D35','D36','D37','D38']
			if bitacora[contador-1]!=0:
				query=db(db.t_bitacora.f_sustancia==ids[contador-1]).select(orderby=db.t_bitacora.f_fechaingreso)
				y=0;
				for i in query:
					if y<24:
						if(i.f_fechaingreso.month==int(mes) and i.f_fechaingreso.year==int(year)):
							ws2[x[y]]= i.f_ingreso/1000
							ws2[x[y]].font= ft3
							ws2[x[y]].alignment = cen
							y=y+1
			x = ['E15','E16','E17','E18','E19','E20','E21','E22','E23','E24','E25','E26','E27','E28','E29','E30','E31','E32','E33','E34','E35','E36','E37','E38']
			if bitacora[contador-1]!=0:
				query=db(db.t_bitacora.f_sustancia==ids[contador-1]).select(orderby=db.t_bitacora.f_fechaingreso)
				y=0;
				for i in query:
					if y<24:
						if(i.f_fechaingreso.month==int(mes) and i.f_fechaingreso.year==int(year)):
							ws2[x[y]]= i.f_consumo/1000
							ws2[x[y]].font= ft3
							ws2[x[y]].alignment = cen
							y=y+1
			x = ['F15','F16','F17','F18','F19','F20','F21','F22','F23','F24','F25','F26','F27','F28','F29','F30','F31','F32','F33','F34','F35','F36','F37','F38']
			if bitacora[contador-1]!=0:
				query=db(db.t_bitacora.f_sustancia==ids[contador-1]).select(orderby=db.t_bitacora.f_fechaingreso)
				y=0;
				for i in query:
					if y<24:
						if(i.f_fechaingreso.month==int(mes) and i.f_fechaingreso.year==int(year)):
							hoja=hoja+1
							ws2[x[y]]= i.f_cantidad/1000
							ws2[x[y]].font= ft3
							ws2[x[y]].alignment = cen
							y=y+1

			#Pie de Pagina
			ws2['A39'] = '1. Los saldos serán reportados en:'
			ws2['A40'] = 'Kgs. Para sustancias en estado sólido ó Lts. Para sustancias en estado líquido, especificando la densidad de la sustancia en el último caso.'		
	wb.save('Reporte Universidad Simon Bolivar.xlsx')
	return locals()

def reporte_bitacora():
	sust = request.vars['sust']
	esp= request.vars['esp']
	wb = Workbook()
	ws2 = wb.active
	cen = Alignment(horizontal='center', vertical='distributed')
	rig = Alignment(horizontal='right')
	lef = Alignment(horizontal='left')
	ft1 = Font(name='Arial', size=10, bold=True)
	ft2 = Font(name='Arial', size=10, bold=False)
	ft3 = Font(name='Arial', size=8)
	ws2.font = ft2
	now = datetime.datetime.now()
	query3=db((db.t_bitacora.f_sustancia==int(sust))&(db.t_bitacora.f_espaciofisico==esp)).select()
	bitacora=0;
	for z in query3:
		bitacora=bitacora+1

	print(bitacora)

	hoja=0;
	while hoja<bitacora:
		wb = Workbook()
		ws2 = wb.active
		cen = Alignment(horizontal='center', vertical='distributed')
		rig = Alignment(horizontal='right')
		lef = Alignment(horizontal='left')
		ft1 = Font(name='Arial', size=10, bold=True)
		ft2 = Font(name='Arial', size=10, bold=False)
		ft3 = Font(name='Arial', size=8)
		ws2.font = ft2
		now = datetime.datetime.now()
		nombre=db(db.t_sustancias.id==int(sust)).select(db.t_sustancias.f_nombre)
		for i in nombre:
			name=i.f_nombre
		ws2 = wb.create_sheet(name)
		#Encabezado
		ws2.title = name


		#tamaño de las columnas
		for i in ['A', 'D', 'E','F','J','K']:
		   ws2.column_dimensions[i].width = 9
		ws2.column_dimensions['B'].width = 9
		ws2.column_dimensions['C'].width = 17.5
		ws2.column_dimensions['H'].width = 9
		ws2.column_dimensions['I'].width = 10.5
		ws2.column_dimensions['G'].width = 11

		#tamaño de las filas
		ws2.row_dimensions[13].height = 13
		for i in range(1,13):
		    ws2.row_dimensions[i].height = 13
		for i in range(14,42):
		    ws2.row_dimensions[i].height = 13


		#All Merges
		ws2.merge_cells(start_row=5,start_column=2,end_row=5,end_column=7)
		ws2.merge_cells(start_row=7,start_column=3,end_row=7,end_column=5)

		#titulos y datos
		z = ['B5', 'G7', 'F8', 'G8', 'H8','B10','B11','B12','F9','G9','H9']
		ws2['B5'] = 'INFORME DE BITACORA'
		ws2['G7'] = 'FECHA'
		ws2['F8'] = 'DIA'
		ws2['G8'] = 'MES'
		ws2['H8'] = 'AÑO'


		for i in range(5):
		    ws2[z[i]].font = ft1
		    ws2[z[i]].alignment = cen

		ws2['B10'] = 'ESPACIO FISICO:'
		ws2['B11'] = 'SUSTANCIA:'
		ws2['B12'] = 'UNIDAD'


		for i in range(5,11):
		    ws2[z[i]].font = ft1
		    ws2[z[i]].alignment = rig

		ws2['F9'] = now.day
		ws2['G9'] = now.month
		ws2['H9'] = now.year

		w = ['C7','C10','C11','A14','B14','C14','D14','E14','F14','G14']

		for i in range(3):
		    ws2[w[i]].font = ft2

		query7=db((db.t_espaciofisico.id==int(esp))).select(db.t_espaciofisico.f_espacio)
		for v in query7:
			ws2['C10'] = v.f_espacio
		ws2['C7'] = 'UNIVERSIDAD SIMON BOLIVAR'
		ws2['C11'] = name.upper()
		query4=db((db.t_bitacora.f_sustancia==int(sust))).select(db.t_bitacora.f_unidad)
		for l in query4:
			ws2['C12']=l.f_unidad 
		ws2['C12'].font = ft2
		ws2['C12'].alignment = cen 			
		
		ws2['A14'] = 'N°'

		ws2['B14'] = 'Fecha'

		ws2['C14'] = 'Descripcion del proceso'

		ws2['D14'] = 'Ingreso'

		ws2['E14'] = 'Consumo'

		ws2['F14'] = 'Cantidad'

		for i in range(3,10):
		    ws2[w[i]].font = ft1
		    ws2[w[i]].alignment = cen

		print("YA ESTOY aqui")

		x = ['A15','A16','A17','A18','A19','A20','A21','A22','A23','A24','A25','A26','A27','A28','A29','A30','A31','A32','A33','A34','A35','A36','A37','A38']
		y = ['01','02','03','04','05','06','07','08','09','10','11','12','13','14','15','16','17','18','19','20','21','22','23','24']
		for i in range(0,24):
			ws2[x[i]] = y[i]
			ws2[x[i]].font = ft2
			ws2[x[i]].alignment = cen

		x = ['B15','B16','B17','B18','B19','B20','B21','B22','B23','B24','B25','B26','B27','B28','B29','B30','B31','B32','B33','B34','B35','B36','B37','B38']
		if bitacora!=0:
			query=db((db.t_bitacora.f_sustancia==int(sust))&(db.t_bitacora.f_espaciofisico==esp)).select(orderby=db.t_bitacora.f_fechaingreso)
			y=0;
			for i in query:
				if y<24:
					ws2[x[y]]= i.f_fechaingreso
					print(i.f_fechaingreso)
					ws2[x[y]].font= ft3
					ws2[x[y]].alignment = cen
					y=y+1

		x = ['C15','C16','C17','C18','C19','C20','C21','C22','C23','C24','C25','C26','C27','C28','C29','C30','C31','C32','C33','C34','AC5','C36','C37','C38']
		if bitacora!=0:
			query=db((db.t_bitacora.f_sustancia==int(sust))&(db.t_bitacora.f_espaciofisico==esp)).select(orderby=db.t_bitacora.f_fechaingreso)
			y=0;
			for i in query:
				if y<24:
					ws2[x[y]]= i.f_proceso
					print(i.f_proceso)
					ws2[x[y]].font= ft3
					ws2[x[y]].alignment = cen
					y=y+1
		x = ['D15','D16','D17','D18','D19','D20','D21','D22','D23','D24','D25','D26','D27','D28','D29','D30','D31','D32','D33','D34','D35','D36','D37','D38']
		if bitacora!=0:
			query=db((db.t_bitacora.f_sustancia==int(sust))&(db.t_bitacora.f_espaciofisico==esp)).select(orderby=db.t_bitacora.f_fechaingreso)
			y=0;
			for i in query:
				if y<24:
					ws2[x[y]]= i.f_ingreso
					print(i.f_ingreso)
					ws2[x[y]].font= ft3
					ws2[x[y]].alignment = cen
					y=y+1
		x = ['E15','E16','E17','E18','E19','E20','E21','E22','E23','E24','E25','E26','E27','E28','E29','E30','E31','E32','E33','E34','E35','E36','E37','E38']
		if bitacora!=0:
			query=db((db.t_bitacora.f_sustancia==int(sust))&(db.t_bitacora.f_espaciofisico==esp)).select(orderby=db.t_bitacora.f_fechaingreso)
			y=0;
			for i in query:
				if y<24:
					ws2[x[y]]= i.f_consumo
					print(i.f_consumo)
					ws2[x[y]].font= ft3
					ws2[x[y]].alignment = cen
					y=y+1
		x = ['F15','F16','F17','F18','F19','F20','F21','F22','F23','F24','F25','F26','F27','F28','F29','F30','F31','F32','F33','F34','F35','F36','F37','F38']
		if bitacora!=0:
			query=db((db.t_bitacora.f_sustancia==int(sust))&(db.t_bitacora.f_espaciofisico==esp)).select(orderby=db.t_bitacora.f_fechaingreso)
			y=0;
			for i in query:
				if y<24:
					hoja=hoja+1
					ws2[x[y]]= i.f_cantidad
					print(i.f_cantidad)
					ws2[x[y]].font= ft3
					ws2[x[y]].alignment = cen
					y=y+1

		#Pie de Pagina
		ws2['A39'] = '1. Los saldos serán reportados en:'
		ws2['A40'] = 'Kgs. Para sustancias en estado sólido ó Lts. Para sustancias en estado líquido, especificando la densidad de la sustancia en el último caso.'		
	wb.save('Reporte Bitacora.xlsx')
	return locals()

def reporte_seccion():
	secc = request.vars['secc']
	seccion = str(db(db.t_seccion.id == int(secc)).select(db.t_seccion.f_seccion))[21:-2]
	lab = db(db.t_seccion.id == secc).select(db.t_seccion.f_laboratorio)
	for h in lab:
		labo=h.f_laboratorio
	query = (db.v_seccion.f_laboratorio == lab)&(db.v_seccion.f_seccion == secc)
	#lab = str(db( db.t_laboratorio.id == lab ).select(db.t_laboratorio.f_nombre))[24:-2]
	wb = Workbook()
	ws2 = wb.active
	cen = Alignment(horizontal='center', vertical='distributed')
	rig = Alignment(horizontal='right')
	lef = Alignment(horizontal='left')
	ft1 = Font(name='Arial', size=10, bold=True)
	ft2 = Font(name='Arial', size=10, bold=False)
	ft3 = Font(name='Arial', size=8)
	ws2.font = ft2
	now = datetime.datetime.now()
	print(labo,secc)
	query3=db((db.v_seccion.f_laboratorio == int(labo))&(db.v_seccion.f_seccion == int(secc))).select()
	bitacora=0;
	for z in query3:
		bitacora=bitacora+1

	print(bitacora)

	hoja=0;
	while hoja<bitacora:
		seccion=unicode(seccion,"utf-8")
		ws2 = wb.create_sheet(seccion)
		#Encabezado
		ws2.title = seccion


		#tamaño de las columnas
		for i in ['A', 'D', 'E','F','J','K']:
		   ws2.column_dimensions[i].width = 9
		ws2.column_dimensions['B'].width = 9
		ws2.column_dimensions['C'].width = 17.5
		ws2.column_dimensions['H'].width = 9
		ws2.column_dimensions['I'].width = 10.5
		ws2.column_dimensions['G'].width = 11

		#tamaño de las filas
		ws2.row_dimensions[13].height = 13
		for i in range(1,13):
		    ws2.row_dimensions[i].height = 13
		for i in range(14,42):
		    ws2.row_dimensions[i].height = 13


		#All Merges
		ws2.merge_cells(start_row=5,start_column=2,end_row=5,end_column=7)
		ws2.merge_cells(start_row=7,start_column=3,end_row=7,end_column=5)

		#titulos y datos
		z = ['B5', 'G7', 'F8', 'G8', 'H8','B10','B11','B12','F9','G9','H9']
		ws2['B5'] = 'INFORME DE SECCION'
		ws2['G7'] = 'FECHA'
		ws2['F8'] = 'DIA'
		ws2['G8'] = 'MES'
		ws2['H8'] = 'AÑO'


		for i in range(5):
		    ws2[z[i]].font = ft1
		    ws2[z[i]].alignment = cen
	    
		ws2['B10'] = 'SECCION:'
		ws2['B11'] = 'LABORATORIO:'


		for i in range(5,11):
		    ws2[z[i]].font = ft1
		    ws2[z[i]].alignment = rig

		ws2['F9'] = now.day
		ws2['G9'] = now.month
		ws2['H9'] = now.year

		w = ['C7','C10','C11','A14','B14','C14','D14','E14','F14','G14']

		for i in range(3):
		    ws2[w[i]].font = ft2

		query7=db((db.t_laboratorio.id==int(labo))).select(db.t_laboratorio.f_nombre)
		for v in query7:
			ws2['C11'] = v.f_nombre
		ws2['C7'] = 'UNIVERSIDAD SIMON BOLIVAR'
		ws2['C10'] = seccion.upper()

		ws2['A14'] = 'N°'

		ws2['B14'] = 'Sustancia'

		ws2['C14'] = 'Cantidad donacion'

		ws2['D14'] = 'Cantidad uso interno'

		ws2['E14'] = 'Total'

		ws2['F14'] = 'Unidad'

		for i in range(3,10):
		    ws2[w[i]].font = ft1
		    ws2[w[i]].alignment = cen

		x = ['A15','A16','A17','A18','A19','A20','A21','A22','A23','A24','A25','A26','A27','A28','A29','A30','A31','A32','A33','A34','A35','A36','A37','A38']
		y = ['01','02','03','04','05','06','07','08','09','10','11','12','13','14','15','16','17','18','19','20','21','22','23','24']
		for i in range(0,24):
			ws2[x[i]] = y[i]
			ws2[x[i]].font = ft2
			ws2[x[i]].alignment = cen

		x = ['B15','B16','B17','B18','B19','B20','B21','B22','B23','B24','B25','B26','B27','B28','B29','B30','B31','B32','B33','B34','B35','B36','B37','B38']
		if bitacora!=0:
			query=db((db.v_seccion.f_laboratorio == int(labo))&(db.v_seccion.f_seccion == int(secc))).select()
			y=0;
			for i in query:
				if y<24:
					ws2[x[y]]= i.f_sustancia
					ws2[x[y]].font= ft3
					ws2[x[y]].alignment = cen
					y=y+1

		x = ['C15','C16','C17','C18','C19','C20','C21','C22','C23','C24','C25','C26','C27','C28','C29','C30','C31','C32','C33','C34','AC5','C36','C37','C38']
		if bitacora!=0:
			query=db((db.v_seccion.f_laboratorio == int(labo))&(db.v_seccion.f_seccion == int(secc))).select()
			y=0;
			for i in query:
				if y<24:
					ws2[x[y]]= i.f_cantidadonacion
					ws2[x[y]].font= ft3
					ws2[x[y]].alignment = cen
					y=y+1
		x = ['D15','D16','D17','D18','D19','D20','D21','D22','D23','D24','D25','D26','D27','D28','D29','D30','D31','D32','D33','D34','D35','D36','D37','D38']
		if bitacora!=0:
			query=db((db.v_seccion.f_laboratorio == int(labo))&(db.v_seccion.f_seccion == int(secc))).select()
			y=0;
			for i in query:
				if y<24:
					ws2[x[y]]= i.f_cantidadusointerno
					ws2[x[y]].font= ft3
					ws2[x[y]].alignment = cen
					y=y+1
		x = ['E15','E16','E17','E18','E19','E20','E21','E22','E23','E24','E25','E26','E27','E28','E29','E30','E31','E32','E33','E34','E35','E36','E37','E38']
		if bitacora!=0:
			query=db((db.v_seccion.f_laboratorio == int(labo))&(db.v_seccion.f_seccion == int(secc))).select()
			y=0;
			for i in query:
				if y<24:
					ws2[x[y]]= i.f_total
					ws2[x[y]].font= ft3
					ws2[x[y]].alignment = cen
					y=y+1
		x = ['F15','F16','F17','F18','F19','F20','F21','F22','F23','F24','F25','F26','F27','F28','F29','F30','F31','F32','F33','F34','F35','F36','F37','F38']
		if bitacora!=0:
			query=db((db.v_seccion.f_laboratorio == int(labo))&(db.v_seccion.f_seccion == int(secc))).select()
			y=0;
			for i in query:
				if y<24:
					hoja=hoja+1
					query
					ws2[x[y]]= i.f_unidad
					ws2[x[y]].font= ft3
					ws2[x[y]].alignment = cen
					y=y+1

		#Pie de Pagina
		ws2['A39'] = '1. Los saldos serán reportados en:'
		ws2['A40'] = 'Kgs. Para sustancias en estado sólido ó Lts. Para sustancias en estado líquido, especificando la densidad de la sustancia en el último caso.'		
	wb.save('Reporte Seccion.xlsx')
	return locals()

def reporte_sust():
	secc = request.vars['secc']
	sust = request.vars['sust']
	seccion = str(db(db.t_seccion.id == int(secc)).select(db.t_seccion.f_seccion))[21:-2]
	lab = db(db.t_seccion.id == secc).select(db.t_seccion.f_laboratorio)
	for h in lab:
		labo=h.f_laboratorio
	query = (db.t_inventario.f_laboratorio == lab)&(db.t_inventario.f_seccion == secc)
	#lab = str(db( db.t_laboratorio.id == lab ).select(db.t_laboratorio.f_nombre))[24:-2]
	wb = Workbook()
	ws2 = wb.active
	cen = Alignment(horizontal='center', vertical='distributed')
	rig = Alignment(horizontal='right')
	lef = Alignment(horizontal='left')
	ft1 = Font(name='Arial', size=10, bold=True)
	ft2 = Font(name='Arial', size=10, bold=False)
	ft3 = Font(name='Arial', size=8)
	ws2.font = ft2
	now = datetime.datetime.now()
	print(labo,secc)
	query3=db((db.t_inventario.f_laboratorio == int(labo))&(db.t_inventario.f_seccion == int(secc))&(db.t_inventario.f_sustancia==int(sust))).select()
	bitacora=0;
	for z in query3:
		bitacora=bitacora+1

	print(bitacora)

	hoja=0;
	while hoja<bitacora:
		seccion=unicode(seccion,"utf-8")
		ws2 = wb.create_sheet(seccion)
		#Encabezado
		ws2.title = seccion


		#tamaño de las columnas
		for i in ['A', 'D', 'E','F','J','K']:
		   ws2.column_dimensions[i].width = 9
		ws2.column_dimensions['B'].width = 9
		ws2.column_dimensions['C'].width = 17.5
		ws2.column_dimensions['H'].width = 9
		ws2.column_dimensions['I'].width = 10.5
		ws2.column_dimensions['G'].width = 11

		#tamaño de las filas
		ws2.row_dimensions[13].height = 13
		for i in range(1,13):
		    ws2.row_dimensions[i].height = 13
		for i in range(14,42):
		    ws2.row_dimensions[i].height = 13


		#All Merges
		ws2.merge_cells(start_row=5,start_column=2,end_row=5,end_column=7)
		ws2.merge_cells(start_row=7,start_column=3,end_row=7,end_column=5)

		#titulos y datos
		z = ['B5', 'G7', 'F8', 'G8', 'H8','B10','B11','B12','F9','G9','H9']
		ws2['B5'] = 'INFORME DE SECCION'
		ws2['G7'] = 'FECHA'
		ws2['F8'] = 'DIA'
		ws2['G8'] = 'MES'
		ws2['H8'] = 'AÑO'


		for i in range(5):
		    ws2[z[i]].font = ft1
		    ws2[z[i]].alignment = cen
	    
		ws2['B10'] = 'SECCION:'
		ws2['B11'] = 'LABORATORIO:'
		ws2['B12'] = 'SUSTANCIA:'


		for i in range(5,11):
		    ws2[z[i]].font = ft1
		    ws2[z[i]].alignment = rig

		ws2['F9'] = now.day
		ws2['G9'] = now.month
		ws2['H9'] = now.year

		w = ['C7','C10','C11','C12','A14','B14','C14','D14','E14','F14','G14']

		for i in range(4):
		    ws2[w[i]].font = ft2

		query7=db((db.t_laboratorio.id==int(labo))).select(db.t_laboratorio.f_nombre)
		for v in query7:
			ws2['C11'] = v.f_nombre
		ws2['C7'] = 'UNIVERSIDAD SIMON BOLIVAR'
		ws2['C10'] = seccion.upper()
		query7=db((db.t_sustancias.id==int(sust))).select(db.t_sustancias.f_nombre)
		for v in query7:
			ws2['C12'] = v.f_nombre

		ws2['A14'] = 'N°'

		ws2['B14'] = 'Sustancia'

		ws2['C14'] = 'Cantidad donacion'

		ws2['D14'] = 'Cantidad uso interno'

		ws2['E14'] = 'Total'

		ws2['F14'] = 'Unidad'

		for i in range(4,11):
		    ws2[w[i]].font = ft1
		    ws2[w[i]].alignment = cen

		x = ['A15','A16','A17','A18','A19','A20','A21','A22','A23','A24','A25','A26','A27','A28','A29','A30','A31','A32','A33','A34','A35','A36','A37','A38']
		y = ['01','02','03','04','05','06','07','08','09','10','11','12','13','14','15','16','17','18','19','20','21','22','23','24']
		for i in range(0,24):
			ws2[x[i]] = y[i]
			ws2[x[i]].font = ft2
			ws2[x[i]].alignment = cen

		x = ['B15','B16','B17','B18','B19','B20','B21','B22','B23','B24','B25','B26','B27','B28','B29','B30','B31','B32','B33','B34','B35','B36','B37','B38']
		if bitacora!=0:
			query=db((db.t_inventario.f_laboratorio == int(labo))&(db.t_inventario.f_seccion == int(secc))&(db.t_inventario.f_sustancia==int(sust))).select(db.t_inventario.f_espaciofisico)
			y=0;
			for i in query:
				if y<24:
					query10=db(db.t_espaciofisico.id==i.f_espaciofisico).select(db.t_espaciofisico.f_direccion)
					for a in query10:
						ws2[x[y]]= a.f_direccion
						ws2[x[y]].font= ft3
						ws2[x[y]].alignment = cen
						y=y+1

		x = ['C15','C16','C17','C18','C19','C20','C21','C22','C23','C24','C25','C26','C27','C28','C29','C30','C31','C32','C33','C34','AC5','C36','C37','C38']
		if bitacora!=0:
			query=db((db.t_inventario.f_laboratorio == int(labo))&(db.t_inventario.f_seccion == int(secc))&(db.t_inventario.f_sustancia==int(sust))).select()
			y=0;
			for i in query:
				if y<24:
					ws2[x[y]]= i.f_cantidadonacion
					ws2[x[y]].font= ft3
					ws2[x[y]].alignment = cen
					y=y+1
		x = ['D15','D16','D17','D18','D19','D20','D21','D22','D23','D24','D25','D26','D27','D28','D29','D30','D31','D32','D33','D34','D35','D36','D37','D38']
		if bitacora!=0:
			query=db((db.t_inventario.f_laboratorio == int(labo))&(db.t_inventario.f_seccion == int(secc))&(db.t_inventario.f_sustancia==int(sust))).select()
			y=0;
			for i in query:
				if y<24:
					ws2[x[y]]= i.f_cantidadusointerno
					ws2[x[y]].font= ft3
					ws2[x[y]].alignment = cen
					y=y+1
		x = ['E15','E16','E17','E18','E19','E20','E21','E22','E23','E24','E25','E26','E27','E28','E29','E30','E31','E32','E33','E34','E35','E36','E37','E38']
		if bitacora!=0:
			query=db((db.t_inventario.f_laboratorio == int(labo))&(db.t_inventario.f_seccion == int(secc))&(db.t_inventario.f_sustancia==int(sust))).select()
			y=0;
			for i in query:
				if y<24:
					ws2[x[y]]= i.f_total
					ws2[x[y]].font= ft3
					ws2[x[y]].alignment = cen
					y=y+1
		x = ['F15','F16','F17','F18','F19','F20','F21','F22','F23','F24','F25','F26','F27','F28','F29','F30','F31','F32','F33','F34','F35','F36','F37','F38']
		if bitacora!=0:
			query=db((db.t_inventario.f_laboratorio == int(labo))&(db.t_inventario.f_seccion == int(secc))&(db.t_inventario.f_sustancia==int(sust))).select()
			y=0;
			for i in query:
				if y<24:
					hoja=hoja+1
					query
					ws2[x[y]]= i.f_unidad
					ws2[x[y]].font= ft3
					ws2[x[y]].alignment = cen
					y=y+1

		#Pie de Pagina
		ws2['A39'] = '1. Los saldos serán reportados en:'
		ws2['A40'] = 'Kgs. Para sustancias en estado sólido ó Lts. Para sustancias en estado líquido, especificando la densidad de la sustancia en el último caso.'		
	wb.save('Reporte Seccion2.xlsx')
	return locals()

def reporte_esp():
	esp = request.vars['esp']
	secc = db(db.t_espaciofisico.id == int(esp)).select(db.t_espaciofisico.f_seccion)
	espf = db(db.t_espaciofisico.id == int(esp)).select(db.t_espaciofisico.f_espacio)
	for h in espf:
		espf=h.f_espacio
	for h in secc:
		secc=h.f_seccion
	lab = db(db.t_seccion.id == secc).select(db.t_seccion.f_laboratorio)
	for h in lab:
		labo=h.f_laboratorio
	query = (db.t_inventario.f_laboratorio == lab)&(db.t_inventario.f_seccion == secc)
	#lab = str(db( db.t_laboratorio.id == lab ).select(db.t_laboratorio.f_nombre))[24:-2]
	wb = Workbook()
	ws2 = wb.active
	cen = Alignment(horizontal='center', vertical='distributed')
	rig = Alignment(horizontal='right')
	lef = Alignment(horizontal='left')
	ft1 = Font(name='Arial', size=10, bold=True)
	ft2 = Font(name='Arial', size=10, bold=False)
	ft3 = Font(name='Arial', size=8)
	ws2.font = ft2
	now = datetime.datetime.now()
	print(labo,secc)
	query3=db((db.t_inventario.f_laboratorio == int(labo))&(db.t_inventario.f_seccion == int(secc))&(db.t_inventario.f_espaciofisico==int(esp))).select()
	bitacora=0;
	for z in query3:
		bitacora=bitacora+1

	print(bitacora)

	hoja=0;
	while hoja<bitacora:
		espf=unicode(espf,"utf-8")
		ws2 = wb.create_sheet(espf)
		#Encabezado
		ws2.title = espf


		#tamaño de las columnas
		for i in ['A', 'D', 'E','F','J','K']:
		   ws2.column_dimensions[i].width = 9
		ws2.column_dimensions['B'].width = 9
		ws2.column_dimensions['C'].width = 17.5
		ws2.column_dimensions['H'].width = 9
		ws2.column_dimensions['I'].width = 10.5
		ws2.column_dimensions['G'].width = 11

		#tamaño de las filas
		ws2.row_dimensions[13].height = 13
		for i in range(1,13):
		    ws2.row_dimensions[i].height = 13
		for i in range(14,42):
		    ws2.row_dimensions[i].height = 13


		#All Merges
		ws2.merge_cells(start_row=5,start_column=2,end_row=5,end_column=7)
		ws2.merge_cells(start_row=7,start_column=3,end_row=7,end_column=5)

		#titulos y datos
		z = ['B5', 'G7', 'F8', 'G8', 'H8','B10','B11','B12','F9','G9','H9']
		ws2['B5'] = 'INFORME DE SECCION'
		ws2['G7'] = 'FECHA'
		ws2['F8'] = 'DIA'
		ws2['G8'] = 'MES'
		ws2['H8'] = 'AÑO'


		for i in range(5):
		    ws2[z[i]].font = ft1
		    ws2[z[i]].alignment = cen
	    
		ws2['B10'] = 'SECCION:'
		ws2['B11'] = 'LABORATORIO:'
		ws2['B12'] = 'ESPACIO FISICO:'


		for i in range(5,11):
		    ws2[z[i]].font = ft1
		    ws2[z[i]].alignment = rig

		ws2['F9'] = now.day
		ws2['G9'] = now.month
		ws2['H9'] = now.year

		w = ['C7','C10','C11','C12','A14','B14','C14','D14','E14','F14','G14']

		for i in range(4):
		    ws2[w[i]].font = ft2

		query7=db((db.t_laboratorio.id==int(labo))).select(db.t_laboratorio.f_nombre)
		for v in query7:
			ws2['C11'] = v.f_nombre
		ws2['C7'] = 'UNIVERSIDAD SIMON BOLIVAR'
		ws2['C12'] = espf.upper()
		query8=db((db.t_seccion.id==secc)).select(db.t_sustancias.f_nombre)
		
		for z in query8:
			ws2['C10'] = v.f_nombre

		ws2['A14'] = 'N°'

		ws2['B14'] = 'Sustancia'

		ws2['C14'] = 'Cantidad donacion'

		ws2['D14'] = 'Cantidad uso interno'

		ws2['E14'] = 'Total'

		ws2['F14'] = 'Unidad'

		for i in range(4,11):
		    ws2[w[i]].font = ft1
		    ws2[w[i]].alignment = cen

		x = ['A15','A16','A17','A18','A19','A20','A21','A22','A23','A24','A25','A26','A27','A28','A29','A30','A31','A32','A33','A34','A35','A36','A37','A38']
		y = ['01','02','03','04','05','06','07','08','09','10','11','12','13','14','15','16','17','18','19','20','21','22','23','24']
		for i in range(0,24):
			ws2[x[i]] = y[i]
			ws2[x[i]].font = ft2
			ws2[x[i]].alignment = cen

		x = ['B15','B16','B17','B18','B19','B20','B21','B22','B23','B24','B25','B26','B27','B28','B29','B30','B31','B32','B33','B34','B35','B36','B37','B38']
		if bitacora!=0:
			query=db((db.t_inventario.f_laboratorio == int(labo))&(db.t_inventario.f_seccion == int(secc))&(db.t_inventario.f_espaciofisico==int(esp))).select(db.t_inventario.f_sustancia)
			y=0;
			for i in query:
				if y<24:
					query10=db(db.t_sustancias.id==i.f_sustancia).select(db.t_sustancias.f_nombre)
					for a in query10:
						ws2[x[y]]= a.f_nombre
						ws2[x[y]].font= ft3
						ws2[x[y]].alignment = cen
						y=y+1

		x = ['C15','C16','C17','C18','C19','C20','C21','C22','C23','C24','C25','C26','C27','C28','C29','C30','C31','C32','C33','C34','AC5','C36','C37','C38']
		if bitacora!=0:
			query=db((db.t_inventario.f_laboratorio == int(labo))&(db.t_inventario.f_seccion == int(secc))&(db.t_inventario.f_espaciofisico==int(esp))).select()
			y=0;
			for i in query:
				if y<24:
					ws2[x[y]]= i.f_cantidadonacion
					ws2[x[y]].font= ft3
					ws2[x[y]].alignment = cen
					y=y+1
		x = ['D15','D16','D17','D18','D19','D20','D21','D22','D23','D24','D25','D26','D27','D28','D29','D30','D31','D32','D33','D34','D35','D36','D37','D38']
		if bitacora!=0:
			query=db((db.t_inventario.f_laboratorio == int(labo))&(db.t_inventario.f_seccion == int(secc))&(db.t_inventario.f_espaciofisico==int(esp))).select()
			y=0;
			for i in query:
				if y<24:
					ws2[x[y]]= i.f_cantidadusointerno
					ws2[x[y]].font= ft3
					ws2[x[y]].alignment = cen
					y=y+1
		x = ['E15','E16','E17','E18','E19','E20','E21','E22','E23','E24','E25','E26','E27','E28','E29','E30','E31','E32','E33','E34','E35','E36','E37','E38']
		if bitacora!=0:
			query=db((db.t_inventario.f_laboratorio == int(labo))&(db.t_inventario.f_seccion == int(secc))&(db.t_inventario.f_espaciofisico==int(esp))).select()
			y=0;
			for i in query:
				if y<24:
					ws2[x[y]]= i.f_total
					ws2[x[y]].font= ft3
					ws2[x[y]].alignment = cen
					y=y+1
		x = ['F15','F16','F17','F18','F19','F20','F21','F22','F23','F24','F25','F26','F27','F28','F29','F30','F31','F32','F33','F34','F35','F36','F37','F38']
		if bitacora!=0:
			query=db((db.t_inventario.f_laboratorio == int(labo))&(db.t_inventario.f_seccion == int(secc))&(db.t_inventario.f_espaciofisico==int(esp))).select()
			y=0;
			for i in query:
				if y<24:
					hoja=hoja+1
					query
					ws2[x[y]]= i.f_unidad
					ws2[x[y]].font= ft3
					ws2[x[y]].alignment = cen
					y=y+1

		#Pie de Pagina
		ws2['A39'] = '1. Los saldos serán reportados en:'
		ws2['A40'] = 'Kgs. Para sustancias en estado sólido ó Lts. Para sustancias en estado líquido, especificando la densidad de la sustancia en el último caso.'		
	wb.save('Reporte Seccion3.xlsx')
	return locals()

def reporte_lab():
	lab = request.vars['lab']
	sust = request.vars['sust']
	#lab = str(db( db.t_laboratorio.id == lab ).select(db.t_laboratorio.f_nombre))[24:-2]
	wb = Workbook()
	ws2 = wb.active
	cen = Alignment(horizontal='center', vertical='distributed')
	rig = Alignment(horizontal='right')
	lef = Alignment(horizontal='left')
	ft1 = Font(name='Arial', size=10, bold=True)
	ft2 = Font(name='Arial', size=10, bold=False)
	ft3 = Font(name='Arial', size=8)
	ws2.font = ft2
	now = datetime.datetime.now()
	query3=db((db.t_inventario.f_laboratorio == int(lab))&(db.t_inventario.f_sustancia == int(sust))).select()
	bitacora=0;
	for z in query3:
		bitacora=bitacora+1

	print(bitacora)
	query7=db((db.t_laboratorio.id==int(lab))).select(db.t_laboratorio.f_nombre)
	for v in query7:
		labo = v.f_nombre
	hoja=0;
	while hoja<bitacora:
		labo=unicode(labo,"utf-8")
		ws2 = wb.create_sheet(labo)
		#Encabezado
		ws2.title = labo


		#tamaño de las columnas
		for i in ['A', 'D', 'E','F','J','K']:
		   ws2.column_dimensions[i].width = 9
		ws2.column_dimensions['B'].width = 9
		ws2.column_dimensions['C'].width = 17.5
		ws2.column_dimensions['H'].width = 9
		ws2.column_dimensions['I'].width = 10.5
		ws2.column_dimensions['G'].width = 11

		#tamaño de las filas
		ws2.row_dimensions[13].height = 13
		for i in range(1,13):
		    ws2.row_dimensions[i].height = 13
		for i in range(14,42):
		    ws2.row_dimensions[i].height = 13


		#All Merges
		ws2.merge_cells(start_row=5,start_column=2,end_row=5,end_column=7)
		ws2.merge_cells(start_row=7,start_column=3,end_row=7,end_column=5)

		#titulos y datos
		z = ['B5', 'G7', 'F8', 'G8', 'H8','B11','B12','F9','G9','H9']
		ws2['B5'] = 'INFORME DE SECCION'
		ws2['G7'] = 'FECHA'
		ws2['F8'] = 'DIA'
		ws2['G8'] = 'MES'
		ws2['H8'] = 'AÑO'


		for i in range(5):
		    ws2[z[i]].font = ft1
		    ws2[z[i]].alignment = cen
	    
		ws2['B11'] = 'LABORATORIO:'
		ws2['B12'] = 'SUSTANCIA:'


		for i in range(5,10):
		    ws2[z[i]].font = ft1
		    ws2[z[i]].alignment = rig

		ws2['F9'] = now.day
		ws2['G9'] = now.month
		ws2['H9'] = now.year

		w = ['C7','C10','C11','C12','A14','B14','C14','D14','E14','F14','G14']

		for i in range(4):
		    ws2[w[i]].font = ft2

		ws2['C11'] = labo
		ws2['C7'] = 'UNIVERSIDAD SIMON BOLIVAR'
		query8=db((db.t_sustancias.id==int(sust))).select(db.t_sustancias.f_nombre)
		for z in query8:
			ws2['C12'] = v.f_nombre

		ws2['A14'] = 'N°'

		ws2['B14'] = 'Sustancia'

		ws2['C14'] = 'Cantidad donacion'

		ws2['D14'] = 'Cantidad uso interno'

		ws2['E14'] = 'Total'

		ws2['F14'] = 'Unidad'

		for i in range(4,11):
		    ws2[w[i]].font = ft1
		    ws2[w[i]].alignment = cen

		x = ['A15','A16','A17','A18','A19','A20','A21','A22','A23','A24','A25','A26','A27','A28','A29','A30','A31','A32','A33','A34','A35','A36','A37','A38']
		y = ['01','02','03','04','05','06','07','08','09','10','11','12','13','14','15','16','17','18','19','20','21','22','23','24']
		for i in range(0,24):
			ws2[x[i]] = y[i]
			ws2[x[i]].font = ft2
			ws2[x[i]].alignment = cen

		x = ['B15','B16','B17','B18','B19','B20','B21','B22','B23','B24','B25','B26','B27','B28','B29','B30','B31','B32','B33','B34','B35','B36','B37','B38']
		if bitacora!=0:
			query=db((db.t_inventario.f_laboratorio == int(lab))&(db.t_inventario.f_sustancia == int(sust))).select(db.t_inventario.f_seccion)
			y=0;
			for i in query:
				if y<24:
					query10=db(db.t_seccion.id==i.f_seccion).select(db.t_seccion.f_seccion)
					for a in query10:
						ws2[x[y]]= a.f_seccion
						ws2[x[y]].font= ft3
						ws2[x[y]].alignment = cen
						y=y+1

		x = ['C15','C16','C17','C18','C19','C20','C21','C22','C23','C24','C25','C26','C27','C28','C29','C30','C31','C32','C33','C34','AC5','C36','C37','C38']
		if bitacora!=0:
			query=db((db.t_inventario.f_laboratorio == int(lab))&(db.t_inventario.f_sustancia == int(sust))).select()
			y=0;
			for i in query:
				if y<24:
					ws2[x[y]]= i.f_cantidadonacion
					ws2[x[y]].font= ft3
					ws2[x[y]].alignment = cen
					y=y+1
		x = ['D15','D16','D17','D18','D19','D20','D21','D22','D23','D24','D25','D26','D27','D28','D29','D30','D31','D32','D33','D34','D35','D36','D37','D38']
		if bitacora!=0:
			query=db((db.t_inventario.f_laboratorio == int(lab))&(db.t_inventario.f_sustancia == int(sust))).select()
			y=0;
			for i in query:
				if y<24:
					ws2[x[y]]= i.f_cantidadusointerno
					ws2[x[y]].font= ft3
					ws2[x[y]].alignment = cen
					y=y+1
		x = ['E15','E16','E17','E18','E19','E20','E21','E22','E23','E24','E25','E26','E27','E28','E29','E30','E31','E32','E33','E34','E35','E36','E37','E38']
		if bitacora!=0:
			query=db((db.t_inventario.f_laboratorio == int(lab))&(db.t_inventario.f_sustancia == int(sust))).select()
			y=0;
			for i in query:
				if y<24:
					ws2[x[y]]= i.f_total
					ws2[x[y]].font= ft3
					ws2[x[y]].alignment = cen
					y=y+1
		x = ['F15','F16','F17','F18','F19','F20','F21','F22','F23','F24','F25','F26','F27','F28','F29','F30','F31','F32','F33','F34','F35','F36','F37','F38']
		if bitacora!=0:
			query=db((db.t_inventario.f_laboratorio == int(lab))&(db.t_inventario.f_sustancia == int(sust))).select()
			y=0;
			for i in query:
				if y<24:
					hoja=hoja+1
					query
					ws2[x[y]]= i.f_unidad
					ws2[x[y]].font= ft3
					ws2[x[y]].alignment = cen
					y=y+1

		#Pie de Pagina
		ws2['A39'] = '1. Los saldos serán reportados en:'
		ws2['A40'] = 'Kgs. Para sustancias en estado sólido ó Lts. Para sustancias en estado líquido, especificando la densidad de la sustancia en el último caso.'		
	wb.save('Reporte lab.xlsx')
	return locals()

def reporte_laboratorio():
	lab = request.vars['lab']
	#lab = str(db( db.t_laboratorio.id == lab ).select(db.t_laboratorio.f_nombre))[24:-2]
	wb = Workbook()
	ws2 = wb.active
	cen = Alignment(horizontal='center', vertical='distributed')
	rig = Alignment(horizontal='right')
	lef = Alignment(horizontal='left')
	ft1 = Font(name='Arial', size=10, bold=True)
	ft2 = Font(name='Arial', size=10, bold=False)
	ft3 = Font(name='Arial', size=8)
	ws2.font = ft2
	now = datetime.datetime.now()
	query3=db((db.v_laboratorio.f_laboratorio == int(lab))).select()
	bitacora=0;
	for z in query3:
		bitacora=bitacora+1

	print(bitacora)
	query7=db((db.t_laboratorio.id==int(lab))).select(db.t_laboratorio.f_nombre)
	for v in query7:
		labo = v.f_nombre
	hoja=0;
	while hoja<bitacora:
		#labo=unicode(labo,"utf-8")
		ws2 = wb.create_sheet(labo)
		#Encabezado
		ws2.title = labo


		#tamaño de las columnas
		for i in ['A', 'D', 'E','F','J','K']:
		   ws2.column_dimensions[i].width = 9
		ws2.column_dimensions['B'].width = 9
		ws2.column_dimensions['C'].width = 17.5
		ws2.column_dimensions['H'].width = 9
		ws2.column_dimensions['I'].width = 10.5
		ws2.column_dimensions['G'].width = 11

		#tamaño de las filas
		ws2.row_dimensions[13].height = 13
		for i in range(1,13):
		    ws2.row_dimensions[i].height = 13
		for i in range(14,42):
		    ws2.row_dimensions[i].height = 13


		#All Merges
		ws2.merge_cells(start_row=5,start_column=2,end_row=5,end_column=7)
		ws2.merge_cells(start_row=7,start_column=3,end_row=7,end_column=5)

		#titulos y datos
		z = ['B5', 'G7', 'F8', 'G8', 'H8','B11','F9','G9','H9']
		ws2['B5'] = 'INFORME DE SECCION'
		ws2['G7'] = 'FECHA'
		ws2['F8'] = 'DIA'
		ws2['G8'] = 'MES'
		ws2['H8'] = 'AÑO'


		for i in range(5):
		    ws2[z[i]].font = ft1
		    ws2[z[i]].alignment = cen
	    
		ws2['B11'] = 'LABORATORIO:'


		for i in range(5,9):
		    ws2[z[i]].font = ft1
		    ws2[z[i]].alignment = rig

		ws2['F9'] = now.day
		ws2['G9'] = now.month
		ws2['H9'] = now.year

		w = ['C7','C10','C11','C12','A14','B14','C14','D14','E14','F14','G14']

		for i in range(4):
		    ws2[w[i]].font = ft2

		ws2['C11'] = labo
		ws2['C7'] = 'UNIVERSIDAD SIMON BOLIVAR'

		ws2['A14'] = 'N°'

		ws2['B14'] = 'Sustancia'

		ws2['C14'] = 'Cantidad donacion'

		ws2['D14'] = 'Cantidad uso interno'

		ws2['E14'] = 'Total'

		ws2['F14'] = 'Unidad'

		for i in range(4,11):
		    ws2[w[i]].font = ft1
		    ws2[w[i]].alignment = cen

		x = ['A15','A16','A17','A18','A19','A20','A21','A22','A23','A24','A25','A26','A27','A28','A29','A30','A31','A32','A33','A34','A35','A36','A37','A38']
		y = ['01','02','03','04','05','06','07','08','09','10','11','12','13','14','15','16','17','18','19','20','21','22','23','24']
		for i in range(0,24):
			ws2[x[i]] = y[i]
			ws2[x[i]].font = ft2
			ws2[x[i]].alignment = cen

		x = ['B15','B16','B17','B18','B19','B20','B21','B22','B23','B24','B25','B26','B27','B28','B29','B30','B31','B32','B33','B34','B35','B36','B37','B38']
		if bitacora!=0:
			query=db((db.v_laboratorio.f_laboratorio == int(lab))).select(db.v_laboratorio.f_sustancia)
			y=0;
			for i in query:
				if y<24:
					ws2[x[y]]= i.f_sustancia
					ws2[x[y]].font= ft3
					ws2[x[y]].alignment = cen
					y=y+1

		x = ['C15','C16','C17','C18','C19','C20','C21','C22','C23','C24','C25','C26','C27','C28','C29','C30','C31','C32','C33','C34','AC5','C36','C37','C38']
		if bitacora!=0:
			query=db((db.v_laboratorio.f_laboratorio == int(lab))).select()
			y=0;
			for i in query:
				if y<24:
					ws2[x[y]]= i.f_cantidadonacion
					ws2[x[y]].font= ft3
					ws2[x[y]].alignment = cen
					y=y+1
		x = ['D15','D16','D17','D18','D19','D20','D21','D22','D23','D24','D25','D26','D27','D28','D29','D30','D31','D32','D33','D34','D35','D36','D37','D38']
		if bitacora!=0:
			query=db((db.v_laboratorio.f_laboratorio == int(lab))).select()
			y=0;
			for i in query:
				if y<24:
					ws2[x[y]]= i.f_cantidadusointerno
					ws2[x[y]].font= ft3
					ws2[x[y]].alignment = cen
					y=y+1
		x = ['E15','E16','E17','E18','E19','E20','E21','E22','E23','E24','E25','E26','E27','E28','E29','E30','E31','E32','E33','E34','E35','E36','E37','E38']
		if bitacora!=0:
			query=db((db.v_laboratorio.f_laboratorio == int(lab))).select()
			y=0;
			for i in query:
				if y<24:
					ws2[x[y]]= i.f_total
					ws2[x[y]].font= ft3
					ws2[x[y]].alignment = cen
					y=y+1
		x = ['F15','F16','F17','F18','F19','F20','F21','F22','F23','F24','F25','F26','F27','F28','F29','F30','F31','F32','F33','F34','F35','F36','F37','F38']
		if bitacora!=0:
			query=db((db.v_laboratorio.f_laboratorio == int(lab))).select()
			y=0;
			for i in query:
				if y<24:
					hoja=hoja+1
					query
					ws2[x[y]]= i.f_unidad
					ws2[x[y]].font= ft3
					ws2[x[y]].alignment = cen
					y=y+1

		#Pie de Pagina
		ws2['A39'] = '1. Los saldos serán reportados en:'
		ws2['A40'] = 'Kgs. Para sustancias en estado sólido ó Lts. Para sustancias en estado líquido, especificando la densidad de la sustancia en el último caso.'		
	wb.save('Reporte laboratorio.xlsx')
	return locals()